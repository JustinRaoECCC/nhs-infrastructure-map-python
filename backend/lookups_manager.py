# backend/lookups_manager.py
# Manages the lookups.xlsx file, ensuring the Locations and AssetTypes sheets exist and providing Excel CRUD for lookups.

import os
import glob
import threading
import random

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font

# ─── Paths & constants ──────────────────────────────────────────────────────
HERE = os.path.dirname(__file__)
DATA_DIR = os.path.abspath(os.path.join(HERE, '..', 'data'))
ASSET_TYPES_DIR = os.path.join(DATA_DIR, 'asset_types')
LOOKUPS_PATH = os.path.join(DATA_DIR, 'lookups.xlsx')
LOCATIONS_DIR   = os.path.join(DATA_DIR, 'locations')

# Random color generator
def _get_random_color() -> str:
    """Return a random hex colour string, e.g. '#3fa4c2'."""
    return '#{:06x}'.format(random.randint(0, 0xFFFFFF))


# ─── Per‑asset‑type locks ────────────────────────────────────────────────────
_lock_dict = {}
def _get_lock(key: str):
    if key not in _lock_dict:
        _lock_dict[key] = threading.Lock()
    return _lock_dict[key]

# ─── Ensure data folder & lookups.xlsx exist and are well‑formed ────────────
def ensure_data_folder():
    os.makedirs(DATA_DIR,        exist_ok=True)
    os.makedirs(ASSET_TYPES_DIR, exist_ok=True)
    os.makedirs(LOCATIONS_DIR,   exist_ok=True)

def ensure_lookups_file():
    """
    Guarantee lookups.xlsx exists with both
    'Locations' and 'AssetTypes' sheets (correct headers).
    """
    ensure_data_folder()

    if not os.path.exists(LOOKUPS_PATH):
        # fresh file: create Companies, Locations, and AssetTypes (now with color!)
        with pd.ExcelWriter(LOOKUPS_PATH, engine='openpyxl') as writer:
            # Companies: “company” + “active”
            pd.DataFrame(columns=['company','active']) \
              .to_excel(writer, sheet_name='Companies', index=False)
            # Locations: “location” + parent “company”
            pd.DataFrame(columns=['location','company']) \
              .to_excel(writer, sheet_name='Locations', index=False)
            # AssetTypes: “asset_type” + parent “location” + “color”
            pd.DataFrame(columns=['asset_type','location','color']) \
              .to_excel(writer, sheet_name='AssetTypes', index=False)
        return


    # patch missing sheets if needed
    wb = load_workbook(LOOKUPS_PATH)

    # — create Companies via pandas so header is bold —
    if 'Companies' not in wb.sheetnames:
        with pd.ExcelWriter(LOOKUPS_PATH, engine='openpyxl', mode='a') as writer:
            pd.DataFrame(columns=['company', 'active']) \
              .to_excel(writer, sheet_name='Companies', index=False)
        # reload so we can patch the other sheets below
        wb = load_workbook(LOOKUPS_PATH)

    # — ensure Locations & AssetTypes exist as before —
    changed = False
    if 'Locations' not in wb.sheetnames:
        ws = wb.create_sheet('Locations')
        ws['A1'] = 'location'
        ws['B1'] = 'company'
        changed = True
    if 'AssetTypes' not in wb.sheetnames:
        ws = wb.create_sheet('AssetTypes')
        ws['A1'] = 'asset_type'
        ws['B1'] = 'location'
        ws['C1'] = 'color'
        changed = True
    # — ensure Custom Weights sheet exists ——
    if 'Custom Weights' not in wb.sheetnames:
        ws = wb.create_sheet('Custom Weights')
        ws['A1'] = 'weight'
        ws['B1'] = 'active'
        # bold headers like other lookup sheets
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        changed = True
    # ═══ Workplan Constants (6th sheet) ════════════════════════════════════
    if 'Workplan Constants' not in wb.sheetnames:
        ws = wb.create_sheet('Workplan Constants')
        # Bold headers
        ws['A1'] = 'Field'
        ws['B1'] = 'Value'
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        # Preset constants
        ws.append(['Yearly Budget', ''])
        ws.append(['O&M Current Split', ''])
        changed = True
    if changed:
        wb.save(LOOKUPS_PATH)


# ─── Core lookup routines ───────────────────────────────────────────────────
def read_lookup_list(sheet_name: str) -> list[str]:
    """
    Read all non‑empty values from column A, rows 2+ of LOOKUPS_PATH[sheet_name].
    If the sheet doesn't exist, create it (with header) and return [].
    """
    wb = load_workbook(LOOKUPS_PATH)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        # match our canonical headers
        if sheet_name == 'Locations':
            ws['A1'] = 'location'
            ws['B1'] = 'company'
        elif sheet_name == 'AssetTypes':
            ws['A1'] = 'asset_type'
            ws['B1'] = 'location'
            ws['C1'] = 'color'
        else:
            # for sheets like "Companies" or "Custom Weights"
            ws['A1'] = sheet_name.slice(0, -1).toLowerCase()
        wb.save(LOOKUPS_PATH)
        return []


    ws = wb[sheet_name]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[0]
        if isinstance(v, str) and v.strip():
            result.append(v.strip())
    return result

def append_to_lookup(sheet_name: str, entry_value: str, second_value: str = None) -> bool:
    """
    Append entry_value (trimmed) to LOOKUPS_PATH[sheet_name], if not already present
    (case‐insensitive). Returns True if added, False if duplicate/empty.
    """
    wb = load_workbook(LOOKUPS_PATH)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws['A1'] = 'LocationName' if sheet_name == 'Locations' else 'AssetTypeName'
    else:
        ws = wb[sheet_name]

    val = (entry_value or '').strip()
    if not val:
        return False

    # check dupes in rows 2+
    for cell in ws['A'][1:]:
        if isinstance(cell.value, str) and cell.value.strip().lower() == val.lower():
            return False

    if sheet_name == 'Companies':
        # second_value now is the exact string we want in column B ("" or "TRUE")
        ws.append([val, second_value])
    elif sheet_name == 'Locations':
        ws.append([val, second_value or ''])
    elif sheet_name == 'AssetTypes':
        # write asset_type, parent_location, AND random colour
        # make sure header row has 3 columns: A=asset_type, B=location, C=color
        color = _get_random_color()
        ws.append([val, second_value or '', color])
    elif sheet_name == 'Custom Weights':
        # entry_value = weight, second_value = 'TRUE' or ''
        ws.append([entry_value, second_value or ''])
        wb.save(LOOKUPS_PATH)
        return True
    else:
        ws.append([val])
        
    wb.save(LOOKUPS_PATH)
    return True

# ─── Public lookup APIs ─────────────────────────────────────────────────────
def get_locations() -> list[str]:
    return read_lookup_list('Locations')

def add_new_location(new_loc: str) -> bool:
    added = append_to_lookup('Locations', new_loc)
    if not added:
        return False
    # create a workbook for this location with one sheet per existing asset-type
    path = os.path.join(LOCATIONS_DIR, f'{new_loc}.xlsx')
    wb = Workbook()
    # leave the default sheet in place (blank) for now;
    # asset-type sheets will be added when Confirm Asset Type is clicked
    wb.save(path)
    return True

def get_asset_types() -> list[str]:
    return read_lookup_list('AssetTypes')

def add_new_asset_type_internal(new_asset_type: str) -> dict:
    """
    In LOOKUPS_PATH[sheet_name], ensure row where Col A == asset_type.
    If missing, append [asset_type, <whatever B was>, <random color>] and save.
    Return the hex color string in Col C.
    """
    lock = _get_lock(new_asset_type)
    with lock:
        name = (new_asset_type or '').strip()
        if not name:
            return {'success': False, 'message': 'Invalid asset type.'}

        # 1) append to the global lookup sheet (col A)
        added = append_to_lookup('AssetTypes', name)
        if not added:
            return {'success': True, 'added': False, 'message': 'Asset type already exists.'}

        # 2) for each location workbook, safely load (or rebuild & retry) then add the sheet
        core_cols = [
            'Station ID','Asset Type','Site Name',
            'Province','Latitude','Longitude',
            'Status'
        ]
        for loc_file in glob.glob(os.path.join(LOCATIONS_DIR, '*.xlsx')):
            wb = None
            # try loading; on any failure, delete & recreate, then reload
            try:
                wb = load_workbook(loc_file)
            except Exception:
                try:
                    os.remove(loc_file)
                except OSError:
                    pass
                basename = os.path.splitext(os.path.basename(loc_file))[0]
                add_new_location(basename)
                try:
                    wb = load_workbook(loc_file)
                except Exception:
                    # still bad? give up on this file
                    continue

            # if the sheet already exists, skip it
            if name in wb.sheetnames:
                continue

            # otherwise, create the new asset‑type sheet
            ws = wb.create_sheet(title=name)
            # if the default 'Sheet' placeholder still exists (and there's >1 sheet), drop it
            if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
                wb.remove(wb['Sheet'])

            # write the header row
            for idx, col in enumerate(core_cols, start=1):
                ws.cell(row=2, column=idx, value=col)
            wb.save(loc_file)

        # 3) done all locations—return success immediately (no exception)
        return {'success': True, 'added': True}

def delete_all_data_files() -> dict:
    # remove lookup file
    for p in glob.glob(os.path.join(DATA_DIR,   '*.xlsx')):
        if os.path.basename(p) == 'lookups.xlsx':
            try: os.remove(p)
            except: pass
    # remove all asset‑type workbooks
    for p in glob.glob(os.path.join(ASSET_TYPES_DIR, '*.xlsx')):
        try: os.remove(p)
        except: pass
    # recreate lookups (but asset_types dir stays)
    ensure_lookups_file()
    return {'success': True}

def get_companies() -> list[str]:
    return read_lookup_list('Companies')

def add_new_company(name: str, active: bool = False) -> bool:
    # active=False → blank; active=True → "TRUE"
    flag = "TRUE" if active else ""
    return append_to_lookup('Companies', name, flag)

def update_lookup_parent(sheet_name: str, entry_value: str, parent_value: str) -> bool:
    """
    In LOOKUPS_PATH[sheet_name], find or append rows of (entry, parent).
    * AssetTypes → fill a blank‐parent row first, then append new rows
      (each with its own random color), never reuse colors.
    * Companies/Locations → do nothing if already present; else update or append.
    """
    wb = load_workbook(LOOKUPS_PATH)
    if sheet_name not in wb.sheetnames:
        return False
    ws = wb[sheet_name]

    entry = entry_value.strip()
    parent = parent_value.strip()

    # ——— AssetTypes: blank‑row‑first, then append new ones, never reuse colors ———
    if sheet_name == 'AssetTypes':
        # find all rows matching this asset_type
        matches = [
            row for row in ws.iter_rows(min_row=2, max_col=3)
            if isinstance(row[0].value, str)
               and row[0].value.strip().lower() == entry.lower()
        ]

        # 1) fill in any blank‐parent row
        for row in matches:
            loc = (row[1].value or '').strip()
            if loc == '':
                row[1].value = parent
                wb.save(LOOKUPS_PATH)
                return True

        # 2) if this exact parent already exists, do nothing
        for row in matches:
            if (row[1].value or '').strip().lower() == parent.lower():
                return False

        # 3) otherwise append a brand‐new row with its own color
        ws.append([entry, parent, _get_random_color()])
        wb.save(LOOKUPS_PATH)
        return True

    # ——— Companies & Locations: update in‑place if found, no‑op if same, else append ———
    for row in ws.iter_rows(min_row=2, max_col=2):
        val = (row[0].value or '').strip()
        if val.lower() == entry.lower():
            # if same parent already, no‑op
            if (row[1].value or '').strip().lower() == parent.lower():
                return False
            # else overwrite
            row[1].value = parent
            wb.save(LOOKUPS_PATH)
            return True

    # not found → append new row
    ws.append([entry, parent])
    wb.save(LOOKUPS_PATH)
    return True


def get_asset_type_color(sheet_name: str, asset_type: str) -> str | None:
    """
    Read LOOKUPS_PATH[sheet_name] for asset_type in Col A,
    return the hex color in Col C (or None if missing sheet/row).
    """
    wb = load_workbook(LOOKUPS_PATH, data_only=True)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, max_col=3):
        val = row[0].value
        if isinstance(val, str) and val.strip().lower() == asset_type.strip().lower():
            return row[2].value
    return None

def get_asset_type_color_for_location(
    sheet_name: str,
    asset_type: str,
    location:   str
) -> str | None:
    """
    Return the hex color for the row matching both asset_type AND location.
    """
    wb = load_workbook(LOOKUPS_PATH, data_only=True)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, max_col=3):
        at   = row[0].value
        loc  = (row[1].value or '').strip()
        col  = row[2].value
        if (
          isinstance(at, str)
          and at.strip().lower() == asset_type.strip().lower()
          and loc.lower() == location.strip().lower()
        ):
            return col
    return None


# ─── Algorithm Parameters (4th sheet) ────────────────────────────────────
def read_algorithm_parameters() -> list[dict]:
    """
    Ensure the 'Algorithm Parameters' sheet exists, then return all rows
    as dicts [{'parameter': str, 'weight': int}, …].
    """
    from openpyxl import load_workbook
    import pandas as pd

    # Create sheet if missing
    if not os.path.exists(LOOKUPS_PATH):
        ensure_lookups_file()
    wb = load_workbook(LOOKUPS_PATH, data_only=True)
    if 'Algorithm Parameters' not in wb.sheetnames:
        # create blank sheet with proper header
        with pd.ExcelWriter(LOOKUPS_PATH, engine='openpyxl', mode='a') as writer:
            pd.DataFrame(columns=['Parameter','Weight']) \
              .to_excel(writer, sheet_name='Algorithm Parameters', index=False)
        wb = load_workbook(LOOKUPS_PATH, data_only=True)

    ws = wb['Algorithm Parameters']
    result = []
    # columns: Applies To, Parameter, Condition, MaxWeight, Option, Weight
    for applies_to, param, condition, max_weight, option, weight, selected in \
        ws.iter_rows(min_row=2, max_col=7, values_only=True):
        if param is None:
            continue
        result.append({
            'applies_to':  str(applies_to or ''),
            'parameter':   str(param),
            'condition':   str(condition or ''),
            'max_weight':  int(max_weight) if max_weight is not None else None,
            'option':      str(option or ''),
            'weight':      int(weight) if weight is not None else None,
            'selected':    True if isinstance(selected, str) and selected.strip().upper() == 'TRUE' else False
        })
    return result

def write_algorithm_parameters(params: list[dict]) -> dict:
    wb = load_workbook(LOOKUPS_PATH)
    if 'Algorithm Parameters' in wb.sheetnames:
        wb.remove(wb['Algorithm Parameters'])
    ws = wb.create_sheet('Algorithm Parameters')
    # write full header
    ws.append(['Applies To','Parameter','Condition','MaxWeight','Option','Weight','Selected'])
    # write every passed row (one for each option)
    for row in params:
        ws.append([
            row['applies_to'],
            row['parameter'],
            row['condition'],
            row['max_weight'],
            row['option'],
            row['weight'],
            'TRUE' if row.get('selected', False) else ''
        ])

    wb.save(LOOKUPS_PATH)
    return {'success': True}


# ─── Workplan Details (5th sheet) ────────────────────────────────────────────
def read_workplan_details() -> list[dict]:
    """
    Ensure 'Workplan Details' sheet exists, then return rows:
    [{'parameter': str, 'value': any}, ...].
    """
    from openpyxl import load_workbook
    import pandas as pd

    # Guarantee lookup file
    ensure_lookups_file()

    wb = load_workbook(LOOKUPS_PATH, data_only=True)
    if 'Workplan Details' not in wb.sheetnames:
        # create blank sheet with header
        with pd.ExcelWriter(LOOKUPS_PATH, engine='openpyxl', mode='a') as writer:
            pd.DataFrame(columns=['Parameter','Value']) \
              .to_excel(writer, sheet_name='Workplan Details', index=False)
        wb = load_workbook(LOOKUPS_PATH, data_only=True)

    ws = wb['Workplan Details']
    result = []
    for param, val in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if param is None:
            continue
        result.append({
            'parameter': str(param),
            'value': val
        })
    return result

def write_workplan_details(entries: list[dict]) -> dict:
    """
    Overwrite the 'Workplan Details' sheet with entries list.
    """
    from openpyxl import load_workbook

    wb = load_workbook(LOOKUPS_PATH)
    # remove old sheet if present
    if 'Workplan Details' in wb.sheetnames:
        wb.remove(wb['Workplan Details'])
    ws = wb.create_sheet('Workplan Details')
    # write header
    ws.append(['Parameter','Value'])
    # write each row
    for e in entries:
        ws.append([e.get('parameter',''), e.get('value','')])
    wb.save(LOOKUPS_PATH)
    return {'success': True}

# ─── Custom Weights APIs ───────────────────────────────────────────────────
def read_custom_weights() -> list[dict]:
    """
    Return list of {'weight': str, 'active': bool}
    """
    wb = load_workbook(LOOKUPS_PATH, data_only=True)
    if 'Custom Weights' not in wb.sheetnames:
        # created earlier in ensure_data_folder
        ensure_lookups_file()
        wb = load_workbook(LOOKUPS_PATH, data_only=True)
    ws = wb['Custom Weights']
    out = []
    for w, flag in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if w is None: continue
        out.append({
            'weight': str(w),
            'active': isinstance(flag, str) and flag.strip().upper() == 'TRUE'
        })
    return out

def add_custom_weight(weight: str, active: bool = False) -> bool:
    """Append weight to Custom Weights sheet; mark active if requested."""
    flag = 'TRUE' if active else ''
    return append_to_lookup('Custom Weights', weight, flag)

# ─── Workplan Constants APIs (6th sheet) ────────────────────────────────────
def read_workplan_constants() -> list[dict]:
    """
    Ensure 'Workplan Constants' exists, then return rows as
    [{'field': str, 'value': any}, …].
    """
    if not os.path.exists(LOOKUPS_PATH):
        ensure_lookups_file()
    wb = load_workbook(LOOKUPS_PATH, data_only=True)
    if 'Workplan Constants' not in wb.sheetnames:
        ensure_lookups_file()
        wb = load_workbook(LOOKUPS_PATH, data_only=True)
    ws = wb['Workplan Constants']
    out = []
    for field, val in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if field is None: continue
        out.append({'field': str(field), 'value': val})
    return out

def write_workplan_constants(entries: list[dict]) -> dict:
    """
    Overwrite the 'Workplan Constants' sheet with the given entries.
    """
    wb = load_workbook(LOOKUPS_PATH)
    if 'Workplan Constants' in wb.sheetnames:
        wb.remove(wb['Workplan Constants'])
    ws = wb.create_sheet('Workplan Constants')
    # Header row
    ws.append(['Field', 'Value'])
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    for e in entries:
        ws.append([e.get('field',''), e.get('value','')])
    wb.save(LOOKUPS_PATH)
    return {'success': True}
