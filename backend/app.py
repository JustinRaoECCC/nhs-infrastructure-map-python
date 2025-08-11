# backend/app.py
# Defines the Eel‑exposed functions that drive the front‑end, delegating all data operations to the DataManager façade.

import os
import glob
import openpyxl
from openpyxl.utils import get_column_letter

import pandas as pd
import eel

from .lookups_manager import (
    ensure_data_folder,
    ensure_lookups_file,
    delete_all_data_files,
    read_custom_weights,
    add_custom_weight as lm_add_custom_weight,
    get_asset_type_color,
    get_asset_type_color_for_location,
    read_algorithm_parameters,
    write_algorithm_parameters,
    read_workplan_details,
    write_workplan_details,
    read_workplan_constants,
    write_workplan_constants
)
from .data_manager     import DataManager
from .data_nuke import data_nuke
from .bulk_importer import get_sheet_names, import_sheet_data
from .repairs_manager import save_repair
from .algorithm import optimize_workplan as _optimize_workplan

# ─── Station file constants ─────────────────────────────────────────────────
HERE = os.path.dirname(__file__)
DATA_DIR = os.path.abspath(os.path.join(HERE, '..', 'data'))
ASSET_TYPES_DIR = os.path.join(DATA_DIR, 'asset_types')

# ─── Bootstrap folder + lookups on import ──────────────────────────────────
ensure_data_folder()
ensure_lookups_file()
dm = DataManager()

# Register excel import endpoints now that dm exists, and inject dm safely
from . import excel_import
excel_import.set_dm(dm)

# ─── Exposed Lookup APIs ───────────────────────────────────────────────────
@eel.expose
def get_locations():
    return dm.get_locations()

@eel.expose
def add_new_location(new_loc):
    return dm.add_location(new_loc)

@eel.expose
def get_asset_types():
    return dm.get_asset_types()

@eel.expose
def add_new_asset_type(new_at):
    return dm.add_asset_type(new_at)

@eel.expose
def get_algorithm_parameters():
    """Return saved algorithm parameters from lookups.xlsx"""
    return read_algorithm_parameters()

@eel.expose
def save_algorithm_parameters(params):
    """
    params: list of {parameter: str, weight: int}
    """
    return write_algorithm_parameters(params)


# ─── Station data APIs ──────────────────────────────────────────────────────
@eel.expose
def get_infrastructure_data():
    stations = dm.list_stations()
    # inject the saved color per province→asset_type
    for stn in stations:
        # 1) try the specific (asset_type + province) match
        # (our Eel‑exposed wrapper now takes exactly two args)
        col = get_asset_type_color_for_location(
                 stn['asset_type'],
                 stn['province']
              )
        # 2) fallback to the generic first‑row lookup
        if not col:
            col = get_asset_type_color('AssetTypes', stn['asset_type'])
        stn['color'] = col or '#000000'

    # 1) Build the union of all extra‑sections per asset_type
    #    schema_map = { asset_type: { section_name: set(fields) } }
    schema_map: dict[str, dict[str,set[str]]] = {}
    for stn in stations:
        at = stn.get('asset_type')
        for key in stn:
            if ' – ' not in key:
                continue
            section, field = key.split(' – ', 1)
            schema_map.setdefault(at, {}).setdefault(section, set()).add(field)

    # 2) Back‑fill each station so it has every section/field in its asset_type’s schema
    for stn in stations:
        at = stn.get('asset_type')
        for section, fields in schema_map.get(at, {}).items():
            for field in fields:
                compound_key = f"{section} – {field}"
                # if this station didn’t have it, give it an empty slot
                if compound_key not in stn:
                    stn[compound_key] = None

    return stations


@eel.expose
def create_new_station(station_obj: dict):
    return dm.create_station(station_obj)

@eel.expose
def create_new_repair(station_id: str, repair_obj: dict):
    """
    Called by your modal’s createNewRepair:
      window.electronAPI.createNewRepair(stationId, { ranking, cost, freq });
    """
    return dm.save_repair(station_id, repair_obj)

@eel.expose
def get_companies():
    return dm.get_companies()

@eel.expose
def add_new_company(name, active=False):
    """
    If active=False: append to Companies with blank active (for the dropdown only).
    If active=True: set the existing row’s active col to “TRUE” (so it shows up in filters).
    """
    from .lookups_manager import update_lookup_parent
    if active:
        ok = update_lookup_parent('Companies', name, 'TRUE')
        return {"success": ok, "added": False}
    else:
        return dm.add_company(name, active)


@eel.expose
def get_active_companies():
    """
    Return only those companies whose “active” column is exactly "TRUE".
    """
    from openpyxl import load_workbook
    from .lookups_manager import LOOKUPS_PATH

    wb = load_workbook(LOOKUPS_PATH, data_only=True)
    if 'Companies' not in wb.sheetnames:
        return []
    ws = wb['Companies']
    out = []
    for name, flag in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if isinstance(name, str) and isinstance(flag, str) \
           and flag.strip().upper() == 'TRUE':
            out.append(name)
    return out

@eel.expose
def get_locations_for_company(company_name):
    return dm.get_locations_for_company(company_name)

@eel.expose
def get_asset_types_for_location(company_name, location_name):
    return dm.get_asset_types_for_location(company_name, location_name)

@eel.expose
def add_asset_type_under_location(asset_type_name, company_name, location_name):
    return dm.add_asset_type_under_location(asset_type_name, company_name, location_name)

@eel.expose
def get_active_filters():
    return dm.get_active_filters()

@eel.expose
def add_location_under_company(location_name, company_name):
    """
    Called by the “Add / Select Location” modal:
      window.electronAPI.addLocationUnderCompany(company, location) 
    """
    return dm.add_location_under_company(location_name, company_name)


@eel.expose
def get_excel_sheet_names(base64_data: str):
    return get_sheet_names(base64_data)

@eel.expose
def import_excel_sheet(b64, sheet, location, asset_type):
    return import_sheet_data(b64, sheet, location, asset_type)

@eel.expose
def get_workplan_details():
    """Return saved workplan entries from lookups.xlsx."""
    return read_workplan_details()

@eel.expose
def get_workplan_constants():
    """Return saved workplan constants from lookups.xlsx."""
    return read_workplan_constants()

@eel.expose
def save_workplan_constants(constants):
    """
    constants: list of {field: str, value: any}
    """
    return write_workplan_constants(constants)

@eel.expose
def get_custom_weights():
    """Return saved custom weights (weight + active)."""
    return read_custom_weights()

@eel.expose
def add_custom_weight(weight, active=False):
    """
    weight: string or number
    active: boolean
    """
    ok = lm_add_custom_weight(str(weight), active)
    return {"success": ok}

@eel.expose
def save_workplan_details(entries):
    """
    entries: list of {parameter: str, value: any}
    """
    return write_workplan_details(entries)

@eel.expose
def save_station_details(station_obj):
    # implement/update in DataManager: apply generalInfo & extraSections,
    # then return {"success":True} or {"success":False,"message":...}
    return dm.update_station(station_obj)

@eel.expose
def delete_station(station_id):
    # implement/delete in DataManager: remove the row from Excel (and DB)
    return dm.delete_station(station_id)

@eel.expose
def list_photos(root_dir: str, include_reports: bool = False):
    """
    Return a pruned directory tree:
      - include only files with allowed extensions
      - include only folders that (recursively) contain at least one allowed file
    """
    import os

    # Allowed file types
    allow_exts = {"png", "jpg", "jpeg", "gif", "bmp"}
    if include_reports:
        allow_exts |= {"pdf", "txt"}

    def _mtime(path):
        try:
            return os.path.getmtime(path)
        except Exception:
            return 0

    def _walk(dir_path):
        """
        Returns (node_dict, has_allowed) where has_allowed means this directory
        or any of its descendants contains an allowed file.
        """
        node = {
            "name": os.path.basename(dir_path),
            "type": "folder",
            "path": dir_path,
            "mtime_ts": _mtime(dir_path),
            "children": []
        }

        # If unreadable, treat as empty
        try:
            entries = sorted(os.listdir(dir_path))
        except Exception:
            return node, False

        has_allowed = False

        for entry in entries:
            full = os.path.join(dir_path, entry)
            if os.path.isdir(full):
                child_node, child_has = _walk(full)
                # Keep the folder only if it (recursively) has an allowed file
                if child_has:
                    node["children"].append(child_node)
                    has_allowed = True or has_allowed
            else:
                # File: include only if extension is allowed
                ext = entry.rsplit(".", 1)[-1].lower() if "." in entry else ""
                if ext in allow_exts:
                    node["children"].append({
                        "name": entry,
                        "type": "file",
                        "path": full,
                        "mtime_ts": _mtime(full)
                    })
                    has_allowed = True

        return node, has_allowed

    tree, _ = _walk(root_dir)
    return tree

@eel.expose
def get_photo_data(path: str) -> str:
    """
    Read the image file at `path` and return a data: URL
    so the frontend can display it.
    """
    import base64, mimetypes
    try:
        ctype, _ = mimetypes.guess_type(path)
        with open(path, "rb") as f:
            data = base64.b64encode(f.read()).decode("ascii")
        return f"data:{ctype or 'application/octet-stream'};base64,{data}"
    except Exception as e:
        # return empty so frontend can handle missing file
        print("get_photo_data error:", e)
        return ""
 
@eel.expose
def get_photo_chunks(path: str) -> list[str]:
    """
    Read the image file at `path` and return its base64 data split into
    ASCII-safe chunks (so no single WebSocket frame exceeds ~16 KB).
    """
    import base64, mimetypes, os
    try:
        # Read and encode
        raw = open(path, "rb").read()
        b64 = base64.b64encode(raw).decode("ascii")
        # Split into 16 000-char chunks
        chunk_size = 16000
        return [b64[i : i + chunk_size] for i in range(0, len(b64), chunk_size)]
    except Exception as e:
        print("get_photo_chunks error:", e)
        return []
    
@eel.expose
def stream_photo(path, uid):
    import base64, os
    b64 = base64.b64encode(open(path,"rb").read()).decode("ascii")
    chunk_size = 16000
    for i in range(0, len(b64), chunk_size):
        eel.receive_photo_chunk(uid, b64[i:i+chunk_size])
    eel.receive_photo_done(uid)
    

# ─── Asset‑Type Color APIs ─────────────────────────────────────────────────
@eel.expose
def get_asset_type_color_lookup(asset_type):
    from .lookups_manager import get_asset_type_color
    # returns hex string or None
    return get_asset_type_color('AssetTypes', asset_type)

@eel.expose
def set_asset_type_color(asset_type, color):
    """
    Update the color for an asset type in lookups.xlsx.
    """
    from .lookups_manager import LOOKUPS_PATH
    from openpyxl import load_workbook
    wb = load_workbook(LOOKUPS_PATH)
    if 'AssetTypes' not in wb.sheetnames:
        return {"success": False, "message": "AssetTypes sheet missing."}
    ws = wb['AssetTypes']
    for row in ws.iter_rows(min_row=2, max_col=3):
        # only match on type
        if isinstance(row[0].value, str) \
          and row[0].value.strip().lower() == asset_type.strip().lower() \
          and ((row[1].value or '').strip() == ''):
            # this is the “generic” row (no location)
            row[2].value = color
            wb.save(LOOKUPS_PATH)
            return {"success": True}
    return {"success": False, "message": f"Asset type '{asset_type}' not found."}

# expose the location‑specific color lookup
@eel.expose
def get_asset_type_color_for_location(asset_type, location):
    from .lookups_manager import get_asset_type_color_for_location
    # sheet_name is always 'AssetTypes'
    return get_asset_type_color_for_location('AssetTypes', asset_type, location)

@eel.expose
def set_asset_type_color_for_location(asset_type, location, color):
    from .lookups_manager import LOOKUPS_PATH
    from openpyxl import load_workbook
    wb = load_workbook(LOOKUPS_PATH)
    ws = wb['AssetTypes']
    for row in ws.iter_rows(min_row=2, max_col=3):
        at = row[0].value
        loc = (row[1].value or '').strip()
        if (
          isinstance(at, str)
          and at.strip().lower() == asset_type.strip().lower()
          and loc.lower() == location.strip().lower()
        ):
            row[2].value = color
            wb.save(LOOKUPS_PATH)
            return {"success": True}
    return {"success": False, "message": f"No row for {asset_type}@{location}"}

@eel.expose
def optimize_workplan(payload=None):
    """
    Called by the front-end when the user clicks “Optimize Workplan”.
    `payload` may include:
      - workplan_rows: list[dict] → rows shown in the Workplan table right now
      - param_overall: { parameter_name: overall_percent } (e.g., 25 for 25%)
    """
    stations = dm.list_stations()
    params   = read_algorithm_parameters()
    consts   = read_workplan_constants()

    workplan_rows = None
    param_overall = None
    if isinstance(payload, dict):
        workplan_rows = payload.get("workplan_rows") or payload.get("rows")
        param_overall = payload.get("param_overall") or payload.get("overall_weights")

    return _optimize_workplan(stations, params, consts, workplan_rows, param_overall)

@eel.expose
def get_repairs(station_id: str):
    from .repairs_manager import list_repairs
    return list_repairs(station_id)

@eel.expose
def delete_repair(station_id: str, row_index: int):
    from .repairs_manager import delete_repair as rm_delete
    return rm_delete(station_id, row_index)

@eel.expose
def delete_dir(path: str) -> dict:
    """
    Recursively delete a directory (or file) at `path`.
    Returns {success: bool, message?: str}.
    """
    import os, shutil, stat

    def _on_rm_error(func, p, exc_info):
        # Handle read-only files on Windows by making them writable then retrying
        try:
            os.chmod(p, stat.S_IWRITE)
            func(p)
        except Exception:
            pass

    try:
        if not os.path.exists(path):
            return {"success": False, "message": "Path not found"}
        if os.path.isdir(path):
            shutil.rmtree(path, onerror=_on_rm_error)
        else:
            os.remove(path)
        return {"success": True}
    except Exception as e:
        print("delete_dir error:", e)
        return {"success": False, "message": str(e)}

@eel.expose
def save_file_from_base64(dest_path: str, b64_data: str) -> dict:
    """
    Create/overwrite `dest_path` with bytes from base64 string (no data: prefix).
    """
    import os, base64
    try:
        os.makedirs(os.path.dirname(dest_path) or ".", exist_ok=True)
        with open(dest_path, "wb") as f:
            f.write(base64.b64decode(b64_data))
        return {"success": True}
    except Exception as e:
        print("save_file_from_base64 error:", e)
        return {"success": False, "message": str(e)}

@eel.expose
def save_files_from_base64(files: list, dest_dir: str) -> dict:
    """
    Write multiple files into dest_dir. Each item: {"name": str, "b64": str}
    """
    import os, base64
    try:
        os.makedirs(dest_dir, exist_ok=True)
        count = 0
        for item in (files or []):
            name = str(item.get("name") or "").strip()
            data = item.get("b64") or ""
            if not name or not data:
                continue
            out = os.path.join(dest_dir, name)
            with open(out, "wb") as f:
                f.write(base64.b64decode(data))
            count += 1
        return {"success": True, "saved": count}
    except Exception as e:
        print("save_files_from_base64 error:", e)
        return {"success": False, "message": str(e)}

@eel.expose
def ensure_dir(path: str) -> dict:
    import os
    try:
        os.makedirs(path, exist_ok=True)
        return {"success": True}
    except Exception as e:
        print("ensure_dir error:", e)
        return {"success": False, "message": str(e)}

@eel.expose
def write_text_file(path: str, text: str) -> dict:
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(text or "")
        return {"success": True}
    except Exception as e:
        print("write_text_file error:", e)
        return {"success": False, "message": str(e)}

@eel.expose
def read_text_file(path: str) -> str:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        print("read_text_file error:", e)
        return ""

@eel.expose
def copy_file(src: str, dest: str) -> dict:
    import shutil, os
    try:
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        shutil.copy2(src, dest)
        return {"success": True}
    except Exception as e:
        print("copy_file error:", e)
        return {"success": False, "message": str(e)}

@eel.expose
def copy_files(src_list: list, dest_dir: str) -> dict:
    import shutil, os
    try:
        os.makedirs(dest_dir, exist_ok=True)
        count = 0
        for s in (src_list or []):
            if not s: continue
            shutil.copy2(s, os.path.join(dest_dir, os.path.basename(s)))
            count += 1
        return {"success": True, "copied": count}
    except Exception as e:
        print("copy_files error:", e)
        return {"success": False, "message": str(e)}

@eel.expose
def open_file_natively(path: str) -> dict:
    import os, subprocess, sys
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)            # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
        return {"success": True}
    except Exception as e:
        print("open_file_natively error:", e)
        return {"success": False, "message": str(e)}


@eel.expose
def import_repairs_excel(b64: str) -> dict:
    """
    Parse an .xlsx (first sheet). Return rows as dicts keyed by the EXACT header
    names found in row 1. No type coercion is applied — values are returned as-is.
    """
    import base64, io
    try:
        raw = base64.b64decode(b64)
        wb  = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws  = wb.active

        # Read header row exactly as written in Excel (trim cell text, keep case)
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {"success": False, "message": "Missing header row."}
        headers = [ (str(c).strip() if c is not None else "") for c in header_row ]
        # Remove trailing empty header cells
        while headers and headers[-1] == "":
            headers.pop()
        if not headers:
            return {"success": False, "message": "Empty header row."}

        def _to_text(v):
            if v is None:
                return None
            s = str(v).strip()
            return s if s != "" else None

        def get_any(i, cast=None):
            """
            Return a value that can be a number or a string.
            If `cast` is provided we attempt it; on failure we keep the original text.
            """
            if i is None:
                return None
            v = row[i]
            if v is None:
                return None
            # Prefer text if it's non-empty
            s = str(v).strip()
            if s == "":
                return None
            if cast is None:
                return s
            try:
                # tolerate thousand-separators for floats
                return cast(s.replace(",", "")) if cast is float else cast(s)
            except Exception:
                return s

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            # Turn the current Excel row into { header: cell_value } exactly.
            record = {}
            # zip stops at the shortest; pad to len(headers)
            cells = list(row) + [None] * max(0, len(headers) - len(row))
            for h, v in zip(headers, cells):
                if not h:
                    continue   # ignore blank header cells
                # Keep the value as-is (number stays number, text stays text like "<2")
                record[h] = v
            # Skip rows that are completely empty under the defined headers
            if any(record.get(h) not in (None, "") for h in headers if h):
                rows.append(record)


        return {"success": True, "rows": rows}
    except Exception as e:
        print("import_repairs_excel error:", e)
        return {"success": False, "message": str(e)}

# ─── App startup ────────────────────────────────────────────────────────────
def main():
    eel.init('frontend')
    eel.start('index.html', size=(1200, 800))

if __name__ == '__main__':
    main()
