# backend/data_manager.py
# Implements the façade that routes reads/writes to whichever BaseRepo implementations are plugged in, handles dual‑writes, and lazy migration.
import openpyxl
from .lookups_manager import LOOKUPS_PATH
from .excel_repo    import ExcelRepo
from .db_repo       import DBRepo
from .persistence   import BaseRepo
from .config         import USE_DATABASE
class DataManager:
    def __init__(
        self,
        excel_provider: BaseRepo = None,
        db_provider:    BaseRepo = None
    ):
        """
        Persistence providers can be injected to support
        Excel, SQL, or any future store without changing logic.
        """
        self.use_db = USE_DATABASE
        self.excel = excel_provider or ExcelRepo()
        self.db    = db_provider    or DBRepo()

    # ─── Helpers ──────────────────────────────────────────────────────────────
    def _migrate_locations(self):
        for name in self.excel.get_locations():
            if not self.db.get_location_by_name(name):
                self.db.add_location(name)

    def _migrate_asset_types(self):
        for name in self.excel.get_asset_types():
            if not self.db.get_asset_type_by_name(name):
                self.db.add_asset_type(name)

    def _migrate_stations(self):
        for rec in self.excel.list_stations():
            if not self.db.get_station_by_id(rec["station_id"]):
                self.db.create_station(rec)

    # ─── Lookups ──────────────────────────────────────────────────────────────
    def get_locations(self):
        # always back‑fill SQL
        self._migrate_locations()
        # return from SQL or Excel
        if self.use_db:
            return [loc.name for loc in self.db.list_locations()]
        return self.excel.get_locations()

    def add_location(self, name: str):
        # ExcelRepo.add_location returns a bool; DBRepo.add_location returns a dict.
        # Wrap the Excel bool in the same dict shape so front‑end always sees {success, message}.
        ok = bool(self.excel.add_location(name))
        res_excel = {
            "success": ok,
            "message": None if ok else "Location was empty or already existed."
        }
        res_db = self.db.add_location(name)
        return res_db if self.use_db else res_excel

    # ─── Asset Types ─────────────────────────────────────────────────────────
    def get_asset_types(self):
        self._migrate_asset_types()
        if self.use_db:
            return [at.name for at in self.db.list_asset_types()]
        return self.excel.get_asset_types()

    def add_asset_type(self, name: str):
        # Both backends return dicts here, so we can just pick one
        res_excel = self.excel.add_asset_type(name)
        res_db    = self.db.add_asset_type(name)
        return res_db if self.use_db else res_excel

    # ─── Stations ─────────────────────────────────────────────────────────────
    def list_stations(self):
        # If you're in SQL mode, first migrate and then read from the DB
        if self.use_db:
            self._migrate_stations()
            out = []
            for s in self.db.list_stations():
                rec = {
                    "station_id": s.station_id,
                    "name":       s.name,
                    "province":   s.province,
                    "lat":        s.lat,
                    "lon":        s.lon,
                    "status":     s.status,
                    "asset_type": s.asset_type.name,
                }
                extra = s.extra_data or {}
                for section_name, fields in extra.items():
                    for field_name, field_value in fields.items():
                        rec[f"{section_name} – {field_name}"] = field_value
                out.append(rec)
            return out

        # Excel‑only mode: skip any SQL migration and just return the flat rows
        return self.excel.list_stations()

    def create_station(self, station_obj: dict):
        self._migrate_asset_types()
        self._migrate_locations()
        x = self.excel.create_station(station_obj)
        d = self.db.create_station(station_obj)
        return d if self.use_db else x

    # ─── Repairs ──────────────────────────────────────────────────────────────
    def save_repair(self, station_id: str, repair_obj: dict):
        x = self.excel.save_repair(station_id, repair_obj)
        d = self.db.save_repair(station_id, repair_obj)
        return d if self.use_db else x

    # ─── Sections ─────────────────────────────────────────────────────────────
    def save_sections(self, station_id: str, sections: dict):
        x = self.excel.save_sections(station_id, sections)
        d = self.db.save_sections(station_id, sections)
        return d if self.use_db else x

    # ─── Company ─────────────────────────────────────────────────────────────
    def get_companies(self):
        if self.use_db:
            return [c.name for c in self.db.list_companies()]
        return self.excel.get_companies()

    def add_company(self, name, active: bool = False):
        excel_res = self.excel.add_company(name, active)
        db_res = self.db.add_company(name)
        return db_res if self.use_db else excel_res
    
    def get_locations_for_company(self, company_name: str):
        if self.use_db:
            return self.db.get_locations_for_company(company_name)
        # Excel mode: read LOOKUPS_PATH → “Locations” sheet, filter by company in column B
        from .lookups_manager import LOOKUPS_PATH
        import openpyxl

        wb = openpyxl.load_workbook(LOOKUPS_PATH, data_only=True)
        if 'Locations' not in wb.sheetnames:
            return []
        ws = wb['Locations']
        result = []
        for loc, comp in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if isinstance(loc, str) and comp == company_name:
                result.append(loc)
        return result


    def get_asset_types_for_location(self, company_name: str, location_name: str):
        if self.use_db:
            return self.db.get_asset_types_for_location(company_name, location_name)
        # Excel mode: list the sheets in the location workbook
        from .lookups_manager import LOCATIONS_DIR
        import os, openpyxl
        path = os.path.join(LOCATIONS_DIR, f'{location_name}.xlsx')
        if not os.path.exists(path):
            return []
        wb = openpyxl.load_workbook(path, read_only=True)
        return wb.sheetnames

    def add_asset_type_under_location(self, asset_type_name: str, company_name: str, location_name: str):
        # --- Excel mode only: record parent and create sheet ---
        from .lookups_manager import update_lookup_parent, add_new_location, LOCATIONS_DIR
        import os, openpyxl

        # 1) mark the parent in lookups.xlsx (col B of AssetTypes)
        update_lookup_parent('AssetTypes', asset_type_name, location_name)

        # 2) ensure location workbook exists
        loc_path = os.path.join(LOCATIONS_DIR, f'{location_name}.xlsx')
        if not os.path.exists(loc_path):
            # if missing, rebuild it
            add_new_location(location_name)

        # 3) open and add a new sheet if needed
        wb = openpyxl.load_workbook(loc_path)
        if asset_type_name not in wb.sheetnames:
            # add the new asset-type sheet
            ws = wb.create_sheet(title=asset_type_name)
            # remove the default blank sheet if it's still there
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            headers = [
                'Station ID','Asset Type','Site Name',
                'Province','Latitude','Longitude',
                'Status'
            ]
            for idx, col in enumerate(headers, start=1):
                ws.cell(row=2, column=idx, value=col)
            wb.save(loc_path)
            return {"success": True, "added": True}
        return {"success": True, "added": False}

    def add_location_under_company(self, location_name: str, company_name: str):
        if self.use_db:
            return self.db.add_location_under_company(location_name, company_name)
        # 1) record the parent in lookups.xlsx (col B of Locations)
        from .lookups_manager import update_lookup_parent, add_new_location

        ok = update_lookup_parent('Locations', location_name, company_name)
        # 2) ensure the physical workbook exists
        add_new_location(location_name)
        return {"success": ok}



    def get_active_filters(self):
        if self.use_db:
            return self.db.get_active_filters()
        else:
            from .lookups_manager import read_lookup_list
            companies = read_lookup_list('Companies')
            return {
                "companies": companies,
                "locations": read_lookup_list('Locations'),
                "asset_types": read_lookup_list('AssetTypes')
            }
        
    def import_stations_from_excel(self, location_name: str, sheet_name: str):
        """
        Read data/locations/{location_name}.xlsx → sheet_name,
        build station_obj for each row, and call create_station().
        """
        from .lookups_manager import LOCATIONS_DIR
        import os, openpyxl

        path = os.path.join(LOCATIONS_DIR, f"{location_name}.xlsx")
        if not os.path.exists(path):
            return {"success": False, "message": f"No workbook for location '{location_name}'"}

        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return {"success": False, "message": f"Sheet '{sheet_name}' not found"}

        ws = wb[sheet_name]
        headers = [c.value for c in ws[2]]
        added = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            rec = dict(zip(headers, row))
            # map general info
            gen = {
                "stationId": str(rec.get("Station ID","")).strip(),
                "siteName":  rec.get("Site Name",""),
                "province":  rec.get("Province",""),
                "latitude":  rec.get("Latitude",0),
                "longitude": rec.get("Longitude",0),
                "status":    rec.get("Status","")
            }
            # extras = all other columns split on “ – ”
            extra = {}
            skip = {"Station ID","Asset Type","Site Name","Province","Latitude","Longitude","Status"}
            for k,v in rec.items():
                if k in skip or not k or " – " not in k:
                    continue
                sec, fld = k.split(" – ",1)
                extra.setdefault(sec, {})[fld] = v

            station_obj = {
                "assetType":     sheet_name,
                "generalInfo":   gen,
                "extraSections": extra
            }
            # reuse your existing create logic (Excel+DB)
            self.create_station(station_obj)
            added += 1

        return {"success": True, "added": added}

    def update_station(self, station_obj: dict):
        # first update Excel, then DB (or vice‑versa)
        res_xl = self.excel.update_station(station_obj)
        res_db = self.db.update_station(station_obj)
        return res_db if self.use_db else res_xl

    def delete_station(self, station_id: str):
        # need to know province+asset_type for Excel delete
        # find the record in the flat Excel list
        recs = self.excel.list_stations()
        target = next((r for r in recs if r["station_id"] == station_id), None)
        if not target:
            return {"success": False, "message": f"Station '{station_id}' not found."}

        res_xl = self.excel.delete_station(target)
        res_db = self.db.delete_station(station_id)
        return res_db if self.use_db else res_xl

    # ─── Merge-only: fill blank existing fields for one station ─────────────
    def merge_fields_for_station(self, station_id: str, col_values: dict):
        """
        Delegate to the active persistence layer(s). This only writes to columns
        that already exist and only if the target cells are blank.
        """
        res_xl = self.excel.merge_fields_for_station(station_id, col_values)
        if self.use_db:
            # Optional: if you later support DB merge semantics, call here too.
            try:
                res_db = self.db.merge_fields_for_station(station_id, col_values)  # type: ignore[attr-defined]
            except Exception:
                res_db = {"success": True, "updated": 0, "checked": res_xl.get("checked", 0)}
            return res_db
        return res_xl
