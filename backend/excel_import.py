# backend/excel_import.py
import base64
import io
from typing import Any, Dict
import re

import eel
import openpyxl
import sys

_dm = None
def set_dm(dm_obj):
    global _dm
    _dm = dm_obj

@eel.expose
def import_multiple_stations(file_b64: str):
    print("\n=== [excel_import] ⚡ import_multiple_stations (merge-only, no create) START ===", flush=True)

    dbg: Dict[str, Any] = {
        "stage": "start",
        "stations_scanned": 0,
        "stations_updated": 0,
        "fields_updated": 0,
        "skipped": 0,
        "errors": [],
        "skip_reasons": []
    }

    try:
        data = base64.b64decode(file_b64)
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        # pick sheet named "List of Stations" (case-insensitive), else first
        ws = None
        for s in wb.worksheets:
            if str(s.title).strip().lower() == "list of stations":
                ws = s
                break
        if ws is None:
            ws = wb.worksheets[0]
        dbg["workbook_sheets"] = [s.title for s in wb.worksheets]
        dbg["active_sheet"] = ws.title
        print(f"[excel_import] ✅ Loaded workbook, using sheet = {ws.title}", flush=True)
    except Exception as e:
        print(f"[excel_import] ❌ Failed to load Excel file: {e}", flush=True)
        return {"success": False, "message": f"Could not read Excel: {e}", "debug": dbg}

    # In read_only mode we must iterate, not index
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        dbg["reason"] = "missing_header_row"
        return {"success": False, "message": "Missing header row.", "debug": dbg}
    headers = [ (str(c).strip() if c is not None else "") for c in first_row ]
    dbg["headers"] = headers
    print(f"[excel_import] 📌 Extracted headers: {headers}", flush=True)

    # Build a case-insensitive index and helpers
    norm = lambda s: re.sub(r"\s+", " ", (s or "").strip().lower())
    header_index = { norm(h): i for i, h in enumerate(headers) if (h or "").strip() != "" }

    def pick(*aliases, contains=None):
        """Return (header_name, index) by alias exact match or substring contains."""
        for a in aliases:
            idx = header_index.get(norm(a))
            if idx is not None:
                return headers[idx], idx
        if contains:
            needle = norm(contains)
            for i, h in enumerate(headers):
                if needle in norm(h):
                    return headers[i], i
        return None, None

    # Required: Station ID
    stn_hdr, stn_idx = pick("Station ID")
    if stn_idx is None:
        dbg["reason"] = "missing_station_id_col"
        print("[excel_import] ❌ 'Station ID' column missing in headers", flush=True)
        return {"success": False, "message": "Missing 'Station ID' column", "debug": dbg}

    # Optional alias for the second col: "Name" or "Site Name" — excluded from merge
    name_hdr,  name_idx  = pick("Site Name", "Name")

    dbg["header_map"] = {
        "Station ID": stn_hdr,
        "Name/Site Name": name_hdr,
    }

    # ── Canonicalization & aliasing so incoming headers match your sheet ─────
    def canon(s: str) -> str:
        import re as _re
        s = _re.sub(r"\s+", " ", (s or "").strip())      # collapse whitespace/newlines
        s = _re.sub(r"\s*\(.*?\)\s*$", "", s)            # drop trailing parenthetical notes
        s = _re.sub(r"\s*/\s*", "/", s)                  # normalize slash spacing
        return s.lower()

    # Map incoming (normalized) names to your exact workbook headers
    ALIAS = {
        "technician": "Site Information – Technician",
        "operating office": "Site Information – Operating Office",
        "pin/pid": "Land Ownership – PIN/PID",
        "regional district": "Land Ownership – Regional District",
        "municipality": "Land Ownership – Municipality",
        "document type/contract (type required - but may not yet have it)": "Land Ownership – Document Type/Contract",
        "tenure/pup no./res number": "Land Ownership – File Number",
        "expiry": "Land Ownership – Expiry",
        # Coordinates string → UTM text column (adjust if you store elsewhere)
        "parcelmap coordinates of station on hydex": "General Information – UTM",
        # Helpful trims for lat/long columns that include inline notes
        "latitude bold = hydex infrastructure specific = estimate ": "Latitude",
        "longitude bold = hydex infrastructure specific = estimate ": "Longitude",
    }

    # Columns we should not try to merge (already part of base/identity)
    SKIP_CANON = {
        canon(stn_hdr or ""),        # Station ID
        canon(name_hdr or ""),       # Name/Site Name
        "asset type", "infrastructure",
        # keep these skipped unless you explicitly want to overwrite blanks
        "latitude", "longitude",
    }


    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        print(f"\n[excel_import] 🔄 Processing Excel row {idx}", flush=True)
        if not row or not any(row):
            print("[excel_import] ⚠️ Row empty, skipping", flush=True)
            dbg["skipped"] += 1
            dbg["skip_reasons"].append({"row": idx, "reason": "Empty row"})
            continue

        # Build a raw dict {header -> value} with original header text
        row_dict = { headers[i]: row[i] for i in range(min(len(headers), len(row))) }
        station_raw = row[stn_idx] if stn_idx is not None and stn_idx < len(row) else None
        station_id = str(station_raw or "").strip()

        if not station_id:
            print("[excel_import] ⚠️ Missing Station ID, skipping row", flush=True)
            dbg["skipped"] += 1
            dbg["skip_reasons"].append({"row": idx, "reason": "Missing Station ID"})
            continue

        if station_id.rstrip("X") != station_id:
            print(f"[excel_import] ⚠️ Skipping station '{station_id}' due to 'X' suffix", flush=True)
            dbg["skipped"] += 1
            dbg["skip_reasons"].append({"row": idx, "station_id": station_id, "reason": "Ends with 'X'"})
            continue

        # Build a merge payload of {canonical header -> value}
        # - exclude Station ID & Name
        # - normalize/alias incoming headers to sheet headers
        # - do NOT introduce new columns (repo enforces this)
        skip_keys_exact = {k for k in [stn_hdr, name_hdr] if k}
        col_values: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h or i >= len(row) or h in skip_keys_exact:
                continue
            val = row[i]
            # Only attempt to merge non-empty incoming values
            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue
            h_norm = canon(h)
            if h_norm in SKIP_CANON:
                continue
            # If we know a canonical mapping, use it
            mapped = ALIAS.get(h_norm)
            if mapped:
                col_values[mapped] = val
                continue
            # Otherwise, if the source is already in "Section – Field" form,
            # keep it as-is and let the repo decide if it exists.
            if "–" in h or "-" in h:
                # Normalize spacing around dash/en-dash to " – "
                parts = re.split(r"\s*[–-]\s*", str(h).strip(), maxsplit=1)
                if len(parts) == 2 and parts[0] and parts[1]:
                    normalized = f"{parts[0].strip()} – {parts[1].strip()}"
                    col_values[normalized] = val
                # else: skip unknown single-field headers (avoid creating columns)


        dbg["stations_scanned"] += 1
        if not col_values:
            # nothing to merge
            continue

        if _dm is None:
            print("[excel_import] ❌ DataManager not injected")
            dbg["reason"] = "dm_not_injected"
            return {"success": False, "message": "DataManager not initialized", "debug": dbg}

        # Merge only into existing blank cells; no new columns; no create
        print(f"[excel_import] → merge_fields_for_station({station_id}) with {len(col_values)} keys", flush=True)
        res = _dm.merge_fields_for_station(station_id, col_values)  # type: ignore[attr-defined]
        print(f"[excel_import] ← merge result: {res}", flush=True)
        if not res.get("success"):
            dbg["errors"].append({"row": idx, "station_id": station_id, "error": res.get("message")})
            dbg["skipped"] += 1
        else:
            updated = int(res.get("updated", 0))
            dbg["fields_updated"] += updated
            if updated > 0:
                dbg["stations_updated"] += 1

    print(f"\n[excel_import] 🎯 Finished. Stations scanned: {dbg['stations_scanned']}, "
          f"stations updated: {dbg['stations_updated']}, fields filled: {dbg['fields_updated']}, "
          f"skipped: {dbg['skipped']}", flush=True)
    return {
        "success": True,
        "message": (
            f"Filled {dbg['fields_updated']} field(s) across {dbg['stations_updated']} station(s). "
            f"Scanned {dbg['stations_scanned']}, skipped {dbg['skipped']}."
        ),
        "debug": dbg
    }
