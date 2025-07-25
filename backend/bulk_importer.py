# backend/bulk_importer.py

import base64
import io
import re
import openpyxl
import eel
import os

from .lookups_manager import LOCATIONS_DIR


def get_sheet_names(base64_data: str):
    """
    Decode a base64‑encoded Excel buffer and return its sheet names.
    """
    raw = base64.b64decode(base64_data)
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    return wb.sheetnames


def import_sheet_data(
    base64_data: str,
    sheet_name: str,
    location_filter: str,
    asset_type_filter: str
) -> dict:
    """
    Decode + load the specified sheet from a base64‑encoded Excel file,
    batch‑append all rows into each province workbook in one go.
    """
    # defer the dm import to avoid the circular‑import
    from .app import dm

    # 1) Decode & load incoming workbook
    raw = base64.b64decode(base64_data)
    wb_in = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    if sheet_name not in wb_in.sheetnames:
        return {"success": False, "message": f"Sheet '{sheet_name}' not found."}

    ws_in = wb_in[sheet_name]

    # 2) Find the header row
    core_headers = {"Station ID", "Station Name", "Latitude", "Longitude"}
    headers = None
    hdr_row = None
    for i, row in enumerate(ws_in.iter_rows(values_only=True), start=1):
        texts = {c for c in row if isinstance(c, str)}
        if core_headers.issubset(texts):
            headers = list(row)
            hdr_row = i
            break

    if headers is None:
        return {"success": False, "message": "Header row not found."}

    # 3) Gather data rows
    data_rows = [
        row for row in ws_in.iter_rows(min_row=hdr_row + 1, values_only=True)
        if row and row[0] is not None
    ]

    total = len(data_rows)
    eel.initImportProgress(total)
    added = 0

    # 4) Batch‑write: one workbook per province
    workbooks = {}  # loc_path -> Workbook
    sheets    = {}  # (loc_path, asset_type) -> Worksheet

    for idx, row_vals in enumerate(data_rows, start=1):
        rec         = dict(zip(headers, row_vals))
        province    = str(location_filter or "").strip()
        asset_type  = str(asset_type_filter or "").strip()
        sid         = str(rec.get("Station ID") or "").strip()

        dm.add_location(province)  # ensures the lookup & workbook exist

        # Prepare the flat row
        base = {
            "Station ID": sid,
            "Asset Type": asset_type,
            "Site Name":  str(rec.get("Station Name") or "").strip(),
            "Province":   province,
            "Latitude":   rec.get("Latitude") or 0,
            "Longitude":  rec.get("Longitude") or 0,
            "Status":     str(rec.get("Status") or "UNKNOWN").strip(),
        }

        # Extras
        extras = {}
        skip_cols = set(base.keys())
        for col, val in rec.items():
            if not col or col in skip_cols:
                continue
            parts = re.split(r"\s*[-–]\s*", col, maxsplit=1)
            if len(parts) == 2:
                sec, fld = parts
                extras.setdefault(sec.strip(), {})[fld.strip()] = val
            else:
                extras.setdefault("Extra Data", {})[col] = val

        full_row = {**base}
        for sec, fields in extras.items():
            for fld, val in fields.items():
                full_row[f"{sec} – {fld}"] = val

        # Load (or reuse) the province workbook
        loc_path = os.path.join(LOCATIONS_DIR, f"{province}.xlsx")
        wb_out = workbooks.get(loc_path)
        if wb_out is None:
            wb_out = openpyxl.load_workbook(loc_path)
            workbooks[loc_path] = wb_out

        # Load (or reuse) the asset‑type sheet
        key    = (loc_path, asset_type)
        ws_out = sheets.get(key)
        if ws_out is None:
            if asset_type not in wb_out.sheetnames:
                ws_out = wb_out.create_sheet(title=asset_type)
                header_list = list(full_row.keys())
                for col_idx, heading in enumerate(header_list, start=1):
                    ws_out.cell(row=2, column=col_idx, value=heading)
            else:
                ws_out = wb_out[asset_type]
                header_list = [c.value for c in ws_out[2]]
            sheets[key] = ws_out
        else:
            header_list = [c.value for c in ws_out[2]]

        # Append the row
        ws_out.append([full_row.get(h) for h in header_list])

        added += 1
        eel.updateImportProgress(idx, total)

    # 5) Save every modified workbook once
    for path, wb in workbooks.items():
        wb.save(path)

    return {"success": True, "added": added}
