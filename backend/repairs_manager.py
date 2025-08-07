# backend/repairs_manager.py
# Handles reading/writing per‑station repair logs into individual Excel files under data/repairs/.

import os
import glob
from openpyxl import load_workbook
import openpyxl
from threading import Lock
from .lookups_manager import LOCATIONS_DIR
from .lookups_manager import REPAIRS_DIR

# Where to store repair files
HERE        = os.path.dirname(__file__)
DATA_DIR    = os.path.abspath(os.path.join(HERE, '..', 'data'))
REPAIRS_DIR = os.path.join(DATA_DIR, 'repairs')

# Ensure directory exists
os.makedirs(REPAIRS_DIR, exist_ok=True)
_lock = Lock()

def _ensure_repair_file(station_id: str) -> str:
    """
    Find the station's location by scanning data/locations/*.xlsx,
    then ensure and return the path to that location's repairs workbook.
    """
    location = None
    for loc_file in glob.glob(os.path.join(LOCATIONS_DIR, '*.xlsx')):
        wb = openpyxl.load_workbook(loc_file, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            headers = [c.value for c in ws[2]]
            if 'Station ID' not in headers:
                continue
            idx = headers.index('Station ID')
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row and row[idx] and str(row[idx]).strip() == station_id:
                    location = os.path.splitext(os.path.basename(loc_file))[0]
                    break
            if location:
                break
        if location:
            break
    if not location:
        # fallback if not found
        location = station_id

    path = os.path.join(REPAIRS_DIR, f'{location}_repairs.xlsx')
    if not os.path.exists(path):
        # create a fresh repairs workbook (will have exactly one sheet: "Sheet")
        from openpyxl import Workbook
        wb_new = Workbook()
        wb_new.save(path)

    return path



def save_repair(station_id: str, repair: dict):
    """
    Append this repair into the sheet named "<SiteName> <StationID>"
    inside the <location>_repairs.xlsx workbook.
    """
    # 1) Ensure the workbook exists
    path = _ensure_repair_file(station_id)
    wb = load_workbook(path)

   # 2) Use only the Station ID (max 31 chars) as the tab name
    sheet_name = station_id[:31]

    # 3) If this station’s sheet doesn’t exist yet, create it *once* and write headers
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append([
            'Site Name',
            'Station Number',
            'Repair Name',
            'Severity Ranking',
            'Priority Ranking',
            'Repair Cost',
            'Category'
        ])
    else:
        ws = wb[sheet_name]

    # append the repair row
    ws.append([
        repair['siteName'],
        station_id,
        repair['name'],
        repair['severity'],
        repair['priority'],
        repair['cost'],
        repair['category']
    ])

    wb.save(path)

def list_repairs(station_id: str) -> list[dict]:
    """
    Read back all the saved repairs for this station.
    Returns a list of dicts keyed by your header names.
    """
    path = _ensure_repair_file(station_id)
    wb = load_workbook(path, data_only=True)
    sheet_name = station_id[:31]
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    # read the header row (values_only gives the raw cell values)
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # skip empty rows
        if not any(row):
            continue
        out.append(dict(zip(headers, row)))
    return out

def delete_repair(station_id: str, row_index: int) -> dict:
    """
    Delete the repair at Excel‐row `row_index` (1-based data rows, excluding header)
    for this station.
    """
    path = _ensure_repair_file(station_id)
    wb = load_workbook(path)
    sheet_name = station_id[:31]
    if sheet_name not in wb.sheetnames:
        return {"success": False, "message": f"No repairs sheet for {station_id}"}
    ws = wb[sheet_name]
    # header is row 1, data starts at row 2:
    excel_row = row_index + 1
    if excel_row < 2 or excel_row > ws.max_row:
        return {"success": False, "message": "Row out of range"}
    ws.delete_rows(excel_row)
    wb.save(path)
    return {"success": True}
